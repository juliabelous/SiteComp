import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime,date,time
import time, os


i=0
typeDict=dict()
typeDict['R']='резистор'
typeDict['L']='индуктивность'
typeDict['C']='конденсатор'
typeDict['MC']='микроконтроллеры'
typeDict['МСХ']='микросxемы'
typeDict['ТР']='транзисторы'
typeDict['ДИОДЫ']='ДИОДЫ'
typeDict['ПР']='ПРОЧЕЕ'


#### XРАНЕНИЕ ПОЛ ####

workUsersBook = './Bases/Users.xlsx'
sheetAllUsers = 'B1'
shUserId = 'A'
shUserName = 'B'
shDate = 'C'
#######################
databaseBook='./base/testDateBase.xlsx'


#######################



#New user
def logUser(chatId, username):
    book = load_workbook(workUsersBook)
    sheet = book.active
        
    #Current num
    num=sheet.max_row
    sheet[sheetAllUsers]=num-1
    curUser=int(num)+1         
        
    #ID writing:
    sheetName='A'+str(curUser)
    sheet[sheetName] = chatId

    #NAME writing:
    curUserName = shUserName + str(curUser)
    sheet[curUserName] = username

    #DATE writing:
    date= time.strftime("%x")
    curUserDate=shDate+str(curUser)
    sheet[curUserDate] = date

    #SAVING
    book.save(workUsersBook)

       

#Check user
def checkUser(chatId):
    book = load_workbook(workUsersBook)
    sheet = book.active
    i=3
    while i<100:
        if chatId == sheet['A'+str(i)].value:
            return True
        else:
            i+=1


#Update
def upgr():
    ids=list()
    book = load_workbook(workUsersBook)
    sheet = book.active
    i=3
    while i<100:
        if not str(sheet['A'+str(i)].value)=='None':
            ids.append(str(sheet['A'+str(i)].value))
            i+=1
        else:
            i+=1
            pass
    return(ids)


#Parsing
def parsing(reqest):
    now=datetime.now()
    now=str(now.replace(microsecond=0, tzinfo=None))
    book=load_workbook(databaseBook)
    lis=book[str(reqest[1])]
    
    if reqest[0]=='Внести':     
        #Проверка наличия
        allPos=lis.max_row
        i=2
        while i<=allPos:
            if str(lis['A'+str(i)].value)==str(reqest[2]):
                if str(lis['B'+str(i)].value)==str(reqest[3]):
                    lis['C'+str(i)] = str(int(lis['C'+str(i)].value)+int(reqest[4]))
                    lis['D'+str(i)] = reqest[6]
                    if not reqest[5]=='Пропустить':
                        os.remove('Photos/'+str(lis['E'+str(i)].value)+'.jpg')
                        lis['E'+str(i)] = reqest[5]
                    lis['F'+str(i)] = str(now)
                    if len(reqest)==8:
                        lis['G'+str(i)] = reqest[7]
                    book.save(databaseBook)
                    os.popen('cd /home/pi/ServerComp/Bases; git add *; git commit -m "'+str(datetime.now())+'"; git push' ) 
                    return(('Изменения добавлены! Остаток '+typeDict[reqest[1]]+' номиналом '+reqest[3] +' в корпусе '+reqest[2]+': '+lis['C'+str(i)].value))
            i+=1

        #Маркировка
        lis['A'+str(lis.max_row+1)] = str(reqest[2])
        #Корпус
        lis['B'+str(lis.max_row)] = str(reqest[3])
        #Кол-во
        lis['C'+str(lis.max_row)] = str(reqest[4])
        #Изменил
        lis['D'+str(lis.max_row)] = str(reqest[6])
        #Фото
        lis['E'+str(lis.max_row)] = str(reqest[5])
        #Изменения  дата
        lis['F'+str(i)] = str(now)
        #Место хранения
        lis['G'+str(i)] = reqest[7]
        os.popen('cd /home/pi/ServerComp/Bases; git add *; git commit -m "'+str(datetime.now())+'"; git push' ) 
        book.save(databaseBook)
        return('Добавлена новая запись!')

    if reqest[0]=='Списать':     
        #Проверка наличия
        allPos=lis.max_row
        i=2
        while i<=allPos:
            if str(lis['A'+str(i)].value)==str(reqest[2]):
                if str(lis['B'+str(i)].value)==str(reqest[3]):
                    if not (int(lis['C'+str(i)].value)-int(reqest[4]))<0 :
                        lis['C'+str(i)] = str(int(lis['C'+str(i)].value)-int(reqest[4]))
                        lis['D'+str(i)] = reqest[6]
                        if not reqest[5]=='Пропустить':
                            lis['E'+str(i)] = reqest[5]
                        lis['F'+str(i)] = str(now)
                        os.popen('cd /home/pi/ServerComp/Bases; git add *; git commit -m "'+str(datetime.now())+'"; git push' ) 
                        book.save(databaseBook)  
                        return('Изменения добавлены! Остаток '+typeDict[reqest[1]]+' номиналом '+reqest[3] +' в корпусе '+reqest[2]+': '+lis['C'+str(i)].value)
                    else:
                        return('Отсутствует нужное кол-во! \nОстаток: '+str(lis['C'+str(i)].value))
            i+=1
        return('Запись не найдена! Попробуйте снова!')
    
    if reqest[0]=='Поиск':
         allPos=lis.max_row
         i=2
         while i<=allPos:
             if str(lis['A'+str(i)].value)==str(reqest[2]):  
                 if str(lis['B'+str(i)].value)==str(reqest[3]):
                     reply=('Результат поиска:\nНаправление поиска-'+reqest[1]+'\nКорпус компонента: '+reqest[2]+'\nМаркировка/номинал: '+
                         reqest[3] + '\nКол-во: '+str(int(lis['C'+str(i)].value))+'\nПоследнее изменение за: @'+str(lis['D'+str(i)].value)+
                            '\nВремя изменения: ' + lis['F'+str(i)].value)
                     if not str((lis['E'+str(i)]).value) == 'Пропустить':
                             photoFile='->'+lis['E'+str(i)].value
                             return(reply+photoFile)
                     return(reply)
             i+=1
         return('Запись не найдена! Попробуйте снова!')
    return('Сбой. Повторите снова!')


def fastParsing(reqest):
    fastReq=reqest.split('+')
    book=load_workbook(databaseBook)
    lis=book[str(fastReq[1])]
    allPos=lis.max_row
    i=2
    cur=1
    replyUser=list()
    outputInfo=list()
    while i<=allPos:
        if  len(str(lis['A'+str(i)].value).split(fastReq[2]))>1:
            replyUser.append(str(cur)+' '+typeDict[fastReq[1]].title()+' '+str(lis['A'+str(i)].value)+' в корпусе '+str(lis['B'+str(i)].value)
                             +' в остатке: '+str(lis['C'+str(i)].value)+'. \nМесто расположения: '+str(lis['G'+str(i)].value))
            outputInfo.append('Списать+'+fastReq[1]+'+'+str(lis['A'+str(i)].value)+'+'+str(lis['B'+str(i)].value))
            cur+=1
            if cur>10:
               replyUser.append("Количество  записей велико! Уточните запрос!\n\n\/\/\/\/\/\/\/\/\/\/\\/\/\/\/\/\/\/\/\/\/\n\n/0 Уточнить запрос!\n\n/\/\/\/\/\/\/\/\/\/\\/\/\/\/\/\/\/\/\/\/\ ")
               
               return replyUser,outputInfo 
            i+=1
        else:
            i+=1
    if len(replyUser) == 0:
        replyUser.append('Записей не найдено!')
    return replyUser,outputInfo 

def fastPhoto(reqest):
    fastReq=reqest
    book=load_workbook(databaseBook)
    lis=book[str(fastReq[1])]
    allPos=lis.max_row
    i=2
    replyUser=list()
    while i<=allPos:
        if str(lis['A'+str(i)].value)==str(reqest[3]):  
                 if str(lis['B'+str(i)].value)==str(reqest[2]):
                    replyUser=lis['E'+str(i)].value
                    return replyUser
                 else:
                    i+=1
        else:
            i+=1
    return 'error'

def qrSearch(reqest):
    reply_user=list()
    infoOut=list()
    lists=typeDict.keys()
    book=load_workbook(databaseBook)
    for captain in lists:
        lis=book[str(captain)]
        allPos=lis.max_row
        i=2
        while i<=allPos:
            if str(lis['G'+str(i)].value)==str(reqest):  
                reply_user.append('/'+str(len(reply_user)+1)+' ) '+typeDict[captain]+
                            ' с маркировкой: ' + str(lis['A'+str(i)].value) +
                                    ' в корпусе ' + str(lis['B'+str(i)].value) +
                                  ' отстаток: ' +  str(lis['C'+str(i)].value) + '\n')
                infoOut.append('Списать+'+str(captain)+'+'+str(lis['A'+str(i)].value)+'+'+str(lis['B'+str(i)].value))
                i+=1
            else:
                i+=1
    if len(reply_user)==0:
        reply_user.append('Не найденно!')
        infoOut.append('Не найденно')
    
    return (reply_user,infoOut)


def superSearch(reqest):
    reply_user=list()
    infoOut=list()
    lists=typeDict.keys()
    book=load_workbook(databaseBook)
    for captain in lists:
        lis=book[str(captain)]
        allPos=lis.max_row
        i=2
        while i<=allPos:
            if  len(str(lis['A'+str(i)].value).split(reqest.upper()))>1:
                reply_user.append(str(len(reply_user)+1)+' ) '+typeDict[captain]+
                            ' с маркировкой: ' + str(lis['A'+str(i)].value) +
                                    ' в корпусе ' + str(lis['B'+str(i)].value) +
                                  ' отстаток: ' +  str(lis['C'+str(i)].value) +'\nРасположение: '+str(lis['G'+str(i)].value))
                infoOut.append('Списать+'+str(captain)+'+'+str(lis['A'+str(i)].value)+'+'+str(lis['B'+str(i)].value))
                i+=1
            else:
                i+=1
    if len(reply_user)==0:
        reply_user.append('Не найденно!')
        infoOut.append('Не найденно')
    
    return (reply_user,infoOut)


def superSearchS(reqest):
    reply_user = list()
    infoOut = list()
    lists = typeDict.keys()
    book = load_workbook(databaseBook)
    for captain in lists:
        lis = book[str(captain)]
        allPos = lis.max_row
        i = 2
        while i <= allPos:
            if len(str(lis['A' + str(i)].value).split(reqest.upper())) > 1:
                reply_user.append(str(len(reply_user) + 1) + ' ) ' + typeDict[captain] +
                                  ' с маркировкой: ' + str(lis['A' + str(i)].value) +
                                  ' в корпусе ' + str(lis['B' + str(i)].value) +
                                  ' отстаток: ' + str(lis['C' + str(i)].value) + '\nРасположение: ' + str(
                    lis['G' + str(i)].value))
                infoOut.append(
                    'Списать+' + str(captain) + '+' + str(lis['A' + str(i)].value) + '+' + str(lis['B' + str(i)].value)+ '+' + str(lis['C' + str(i)].value))
                i += 1
            else:
                i += 1
    if len(reply_user) == 0:
        reply_user.append('Не найденно!')
        infoOut.append('Не найденно')

    return (reply_user, infoOut)

def allTypes(Type):
    typesComp=dict()
    typesComp['R']='R'
    typesComp['C']='C'
    typesComp['L']='L'
    typesComp['MC']='MC'
    typesComp['MSH']='МСХ'
    typesComp['TR']='ТР'
    typesComp['D']='ДИОДЫ'
    typesComp['O']= 'ПР'
    infoOut = list()
    book = load_workbook(databaseBook)
    lis = book[typesComp[str(Type)]]
    allPos = lis.max_row
    i = 2
    while i <= allPos:
            infoOut.append(
                'Списать+' + str(Type) + '+' + str(lis['A' + str(i)].value) + '+' + str(
                    lis['B' + str(i)].value) + '+' + str(lis['C' + str(i)].value))
            i+=1
    return infoOut

def outPssComp(name, body, typeEdit):
    infoOut = list()
    lists = typeDict.keys()
    book = load_workbook(databaseBook)
    for captain in lists:
        lis = book[str(captain)]
        allPos = lis.max_row
        i = 2
        while i <= allPos:
            if len((str(lis['A' + str(i)].value).upper()).split(name.upper())) > 1:
                if str(lis['B'+str(i)].value)==str(body):
                    if typeEdit=='output':
                        if int(lis['C' + str(i)].value) == 0:
                            return 'Empty'
                        lis['C' + str(i)]=int(lis['C'+str(i)].value)-1
                        book.save(databaseBook)
                        infoOut.append(
                            'Списать+' + str(captain) + '+' + str(lis['A' + str(i)].value) + '+' + str(
                                lis['B' + str(i)].value) + '+' + str(lis['C' + str(i)].value))
                    if typeEdit == 'input':
                        lis['C' + str(i)] = int(lis['C' + str(i)].value)+1
                        book.save(databaseBook)
                        infoOut.append(
                            'Внести+' + str(captain) + '+' + str(lis['A' + str(i)].value) + '+' + str(
                                lis['B' + str(i)].value) + '+' + str(lis['C' + str(i)].value))

                i += 1
            else:
                i += 1
    if len(infoOut) == 0:
        infoOut.append('Не найденно')
    return (infoOut)

